from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook

from rumydata.field import Integer, Field, Text
from rumydata.rules.column import Unique
from rumydata.table import Layout, CsvFile, ExcelFile, _BaseFile
from rumydata import exception as ex
from tests.utils import mock_no_module
from rumydata import rules


@pytest.fixture
def wb_sheets(tmpdir):
    wb = Workbook()
    ws0 = wb.active
    ws0.append([''])
    ws1 = wb.create_sheet('wrong')
    ws1.append(['x'])
    ws1.append(['aa'])
    ws2 = wb.create_sheet('right')
    ws2.append(['x'])
    ws2.append(['a'])
    p = Path(tmpdir, f"{uuid4().hex}.xlsx")
    wb.save(p)
    yield p


def test_exception_message_structure(tmpdir):
    """
    Exception message structure

    This is a very hokey test to check the behavior of the exception message
    structure. This should be improved. This is going to break all the time as
    rule messages and various minor details get changed in the underlying
    classes.
    """
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(["c1,c2", "1,23", "1,1"]))
    msg = [
        f" - File: {hid}",
        "   - Row: 2",
        "     - Cell: 2,2 (c2)",
        "       - MaxDigit: must have no more than 1 digit characters",
        "   - Column: 1 (c1)",
        "     - Unique: values must be unique",
    ]
    msg = '\n'.join(msg)
    lay = Layout({'c1': Integer(1, rules=[Unique()]), 'c2': Integer(1)})
    try:
        CsvFile(lay).check(p)
    except AssertionError as ae:
        print(ae)
        print(msg)
        assert str(ae).endswith(msg)


def test_documentation():
    """ An equally silly test of documentation output """
    layout = Layout({'x': Field()})
    expected = ' - **x**\n   - cannot be empty/blank'
    assert layout.documentation() == expected


@pytest.mark.parametrize('choice, valid', [
    ('md', True),
    ('html', True),
    ('yyz', False)
])
def test_documentation_types(choice, valid):
    layout = Layout({'x': Field()})
    if valid:
        assert isinstance(layout.documentation(doc_type=choice), str)
    else:
        with pytest.raises(TypeError):
            layout.documentation(doc_type=choice)


@pytest.mark.parametrize('choice, valid', [
    ('md', True),
    ('html', True),
    ('yyz', False)
])
@pytest.mark.parametrize('valid_file', [False, True])
def test_file_output_types(choice, valid, valid_file, tmpdir):
    layout = Layout({'x': Integer(1)})
    p = Path(tmpdir, 'test_file_output_types.csv')
    if valid_file:
        p.write_text('x\n1\n')
    else:
        p.write_text('x\nx\n')

    if valid:
        assert isinstance(CsvFile(layout).check(p, choice), str)
    else:
        with pytest.raises(TypeError):
            CsvFile(layout).check(p, choice)


def test_no_excel(mocker):
    mocker.patch('builtins.__import__', wraps=__import__, side_effect=mock_no_module('openpyxl'))
    with pytest.raises(ModuleNotFoundError):
        ExcelFile(Layout({'x': Integer(1)}))


def test_base_file_stubs():
    assert not _BaseFile(Layout({'x': Integer(1)}))._rows(Path('x'))


def test_excel_wrong_sheet(wb_sheets):
    lay = Layout({'x': Text(1)})
    with pytest.raises(AssertionError):
        ExcelFile(lay, sheet='wrong').check(wb_sheets)


def test_excel_right_sheet(wb_sheets):
    lay = Layout({'x': Text(1)})
    assert not ExcelFile(lay, sheet='right').check(wb_sheets)


def test_file_name_match(tmpdir):
    mock_file = Path(tmpdir, '12345_test_file_report.csv')
    pattern = r'\d{5}_\D*_\D*_report.csv'
    layout = Layout({'x': Integer(1)})
    mock_file.write_text('x\n1\n')
    assert not CsvFile(layout, file_name_pattern=pattern).check(mock_file)


def test_excel_cell_format():
    lay = Layout({'col_a': Text(1)}, use_excel_cell_format=True)
    try:
        lay.check_row([''])
    except AssertionError as e:
        assert 'A0' in str(e)


def test_no_header_true_bad(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['aa,b', 'c,d']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, no_header=True)
    assert False if CsvFile(layout)._list_errors(p) == [None] else True


def test_no_header_true_bad_plus(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['aa,b', 'cc,d']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, no_header=True)
    errors = CsvFile(layout)._list_errors(p)
    assert True if len([x for x in errors if type(x) == ex.RowError]) == 2 else False


def test_no_header_true_good(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['a,b', 'c,d']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, no_header=True)
    assert False if CsvFile(layout)._list_errors(p) != [None] else True


def test_no_header_false_bad(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['a,b', 'c,d']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, no_header=False)
    assert False if CsvFile(layout)._list_errors(p) == [None] else True


def test_no_header_false_good(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['c1,c2', 'a,b']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, no_header=False)
    assert False if CsvFile(layout)._list_errors(p) != [None] else True


def test_no_header_default_good(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['c1,c2', 'a,b']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)})
    assert False if CsvFile(layout)._list_errors(p) != [None] else True


def test_no_header_default_bad(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['a,b', 'c,d']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)})
    assert False if CsvFile(layout)._list_errors(p) == [None] else True


def test_no_header_with_column_rule(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['a', 'a']))
    layout = Layout({'c1': Text(1, rules=[Unique()])}, no_header=True)
    assert True if CsvFile(layout)._has_error(p, Unique.rule_exception()) else False


def test_no_header_with_skip_rows(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['aa', 'aa', 'aa']))
    layout = Layout({'c1': Text(1, rules=[Unique()])}, no_header=True)
    errors = CsvFile(layout, skip_rows=1)._list_errors(p)
    assert True if all([len([x for x in errors if type(x) == ex.ColumnError]) == 1,
                        len([x for x in errors if type(x) == ex.RowError]) == 2]) else False


def test_skip_rows_bad_header(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['', 'c2', 'aa', 'aa']))
    layout = Layout({'c1': Text(2, rules=[Unique()])})
    csv = CsvFile(layout, skip_rows=1)
    assert all(
        [csv._has_error(p, rules.header.ColumnOrder.rule_exception()), not csv._has_error(p, Unique.rule_exception())])


@pytest.mark.parametrize('skip, expected', [
    (1, True),
    (2, False)
])
def test_skip_rows_skips_columns_errors(tmpdir, skip, expected):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['a', 'a', 'a', 'a']))
    layout = Layout({'a': Text(2, rules=[Unique()])})
    csv = CsvFile(layout, skip_rows=skip)
    assert csv._has_error(p, Unique.rule_exception()) is expected


def test_ignore_if_single_good(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['c1,c2', 'x,', 'c,d']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, empty_row_ok=True)
    results = CsvFile(layout, ignore_exceptions={'c1': 'x'}).check(p)
    assert True if not results else False


def test_ignore_if_single_bad(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['c1,c2', 'z,', 'c,d']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, empty_row_ok=True)
    assert False if CsvFile(layout, ignore_exceptions={'c1': 'x'})._list_errors(p) == [None] else True


def test_ignore_if_list(tmpdir):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['c1,c2', 'x,z', 'c,d', ',', 'z,', 'x,']))
    layout = Layout({'c1': Text(1), 'c2': Text(1)}, empty_row_ok=True)
    results = CsvFile(layout, ignore_exceptions={'c1': ['x', 'z']}).check(p)
    assert True if not results else False


@pytest.mark.parametrize("write, read, expect_pass", [
    ('utf-8', 'utf-8', True),
    ('utf-8-sig', 'utf-8-sig', True),
    ('utf-8', 'utf-8-sig', True),
    ('utf-8-sig', 'utf-8', False),
])
def test_bom_sig(tmpdir, write, read, expect_pass):
    hid = uuid4().hex[:5]
    p = Path(tmpdir, hid)
    p.write_text('\n'.join(['column', 'data']), encoding=write)
    layout = Layout({'column': Text(4)})
    cf = CsvFile(layout, encoding=read)

    if expect_pass is True:
        assert not cf.check(p)
    else:
        with pytest.raises(AssertionError):
            cf.check(p)
